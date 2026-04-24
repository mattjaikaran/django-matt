"""
Export model data to various formats (CSV, JSON, JSONL, XLSX).

Usage:
    python manage.py matt_export myapp.Product --format csv --output products.csv
    python manage.py matt_export myapp.Product --format json --output products.json
    python manage.py matt_export myapp.Product --format csv --filter "status=active,created_at__gte=2026-01-01"
    python manage.py matt_export myapp.Product --format csv --fields "id,name,price,category__name"
    python manage.py matt_export myapp.Product --format csv  # stdout if no --output
"""

from __future__ import annotations

import csv
import io
import sys
from typing import Any

from django.apps import apps
from django.core.management.base import CommandError
from django.db import models

import orjson

from django_matt.cli import MattCommand


def _resolve_model(model_path: str) -> type[models.Model]:
    """Resolve 'app_label.ModelName' to a Django model class."""
    parts = model_path.rsplit(".", 1)
    if len(parts) != 2:
        raise CommandError(f"Invalid model path '{model_path}'. Use 'app_label.ModelName' format.")
    app_label, model_name = parts
    try:
        return apps.get_model(app_label, model_name)
    except LookupError:
        raise CommandError(f"Model '{model_path}' not found.")


def _parse_filters(filter_str: str) -> dict[str, Any]:
    """Parse comma-separated Django ORM lookups into a dict."""
    filters: dict[str, Any] = {}
    for part in filter_str.split(","):
        part = part.strip()
        if "=" not in part:
            raise CommandError(f"Invalid filter: '{part}'. Expected 'field=value'.")
        key, value = part.split("=", 1)
        # coerce booleans and None
        if value.lower() == "true":
            value = True
        elif value.lower() == "false":
            value = False
        elif value.lower() == "none":
            value = None
        filters[key.strip()] = value
    return filters


def _detect_related_fields(
    model: type[models.Model], field_names: list[str]
) -> tuple[list[str], list[str]]:
    """Detect FK and M2M fields for queryset optimization."""
    select = []
    prefetch = []
    meta = model._meta
    for name in field_names:
        base = name.split("__")[0]
        try:
            field = meta.get_field(base)
        except Exception:
            continue
        if isinstance(field, models.ForeignKey):
            select.append(base)
        elif isinstance(field, (models.ManyToManyField, models.ManyToManyRel)):
            prefetch.append(base)
    return select, prefetch


def _get_value(obj: Any, field_path: str) -> Any:
    """Traverse dotted/dunder field paths on an object."""
    current = obj
    for part in field_path.split("__"):
        if current is None:
            return None
        current = getattr(current, part, None)
    return current


class Command(MattCommand):
    """Export model data to CSV, JSON, JSONL, or XLSX with filtering and field selection."""

    help = "Export model data to CSV, JSON, JSONL, or XLSX"

    def add_arguments(self, parser):
        super().add_arguments(parser)
        parser.add_argument("model", help="Model path (e.g. myapp.Product)")
        parser.add_argument(
            "--format",
            "-f",
            default="csv",
            choices=["csv", "json", "jsonl", "xlsx"],
            help="Output format (default: csv)",
        )
        parser.add_argument("--output", "-o", help="Output file path (stdout if omitted)")
        parser.add_argument(
            "--filter", help="Comma-separated ORM lookups (e.g. status=active,price__gte=10)"
        )
        parser.add_argument("--fields", help="Comma-separated field names to include")
        parser.add_argument("--exclude", help="Comma-separated field names to exclude")
        parser.add_argument("--limit", type=int, help="Maximum number of rows")
        parser.add_argument("--order-by", help="Comma-separated ordering fields")

    def handle(self, *args, **options):
        model = _resolve_model(options["model"])
        fmt = options["format"]
        output_path = options.get("output")

        # determine fields
        if options.get("fields"):
            field_names = [f.strip() for f in options["fields"].split(",")]
        else:
            field_names = [f.name for f in model._meta.fields]

        # apply exclusions
        if options.get("exclude"):
            excludes = {f.strip() for f in options["exclude"].split(",")}
            field_names = [f for f in field_names if f not in excludes]

        # build queryset
        qs = model.objects.all()

        if options.get("filter"):
            filters = _parse_filters(options["filter"])
            qs = qs.filter(**filters)

        if options.get("order_by"):
            ordering = [f.strip() for f in options["order_by"].split(",")]
            qs = qs.order_by(*ordering)

        # optimize related lookups
        select, prefetch = _detect_related_fields(model, field_names)
        if select:
            qs = qs.select_related(*select)
        if prefetch:
            qs = qs.prefetch_related(*prefetch)

        if options.get("limit"):
            qs = qs[: options["limit"]]

        # xlsx requires output file
        if fmt == "xlsx" and not output_path:
            raise CommandError("--output is required for xlsx format.")

        total = qs.count() if not options.get("limit") else min(qs.count(), options["limit"])
        self.console.header("Data Export", f"{options['model']} -> {fmt.upper()}")
        self.console.info(f"Exporting {total} rows, {len(field_names)} fields")

        if fmt == "csv":
            self._export_csv(qs, field_names, output_path, total)
        elif fmt == "json":
            self._export_json(qs, field_names, output_path, total)
        elif fmt == "jsonl":
            self._export_jsonl(qs, field_names, output_path, total)
        elif fmt == "xlsx":
            self._export_xlsx(qs, field_names, output_path, total)

        if output_path:
            self.console.success(f"Exported {total} rows to {output_path}")

    # ------------------------------------------------------------------
    # Format writers
    # ------------------------------------------------------------------

    def _iter_rows(
        self,
        qs: models.QuerySet,
        field_names: list[str],
        total: int,
    ):
        """Yield dicts for each row with Rich progress."""
        with self.console.progress("Exporting...", total=total) as progress:
            task = progress.add_task("Exporting", total=total)
            for obj in qs.iterator(chunk_size=2000):
                row = {}
                for name in field_names:
                    val = _get_value(obj, name)
                    row[name] = self._serialize_value(val)
                progress.advance(task)
                yield row

    @staticmethod
    def _serialize_value(val: Any) -> Any:
        """Convert a Python value to an export-safe representation."""
        if val is None:
            return ""
        if isinstance(val, models.Model):
            return str(val.pk)
        if hasattr(val, "isoformat"):
            return val.isoformat()
        return val

    def _export_csv(
        self,
        qs: models.QuerySet,
        field_names: list[str],
        output_path: str | None,
        total: int,
    ) -> None:
        buf = io.StringIO() if not output_path else None
        fh = open(output_path, "w", newline="") if output_path else buf  # noqa: SIM115
        try:
            writer = csv.DictWriter(fh, fieldnames=field_names)
            writer.writeheader()
            for row in self._iter_rows(qs, field_names, total):
                writer.writerow(row)
        finally:
            if output_path:
                fh.close()

        if not output_path:
            sys.stdout.write(buf.getvalue())

    def _export_json(
        self,
        qs: models.QuerySet,
        field_names: list[str],
        output_path: str | None,
        total: int,
    ) -> None:
        rows = list(self._iter_rows(qs, field_names, total))
        data = orjson.dumps(rows, option=orjson.OPT_INDENT_2)
        if output_path:
            with open(output_path, "wb") as fh:
                fh.write(data)
        else:
            sys.stdout.write(data.decode())
            sys.stdout.write("\n")

    def _export_jsonl(
        self,
        qs: models.QuerySet,
        field_names: list[str],
        output_path: str | None,
        total: int,
    ) -> None:
        if output_path:
            with open(output_path, "wb") as fh:
                for row in self._iter_rows(qs, field_names, total):
                    fh.write(orjson.dumps(row))
                    fh.write(b"\n")
        else:
            for row in self._iter_rows(qs, field_names, total):
                sys.stdout.write(orjson.dumps(row).decode())
                sys.stdout.write("\n")

    def _export_xlsx(
        self,
        qs: models.QuerySet,
        field_names: list[str],
        output_path: str,
        total: int,
    ) -> None:
        try:
            import openpyxl
        except ImportError:
            raise CommandError(
                "openpyxl is required for xlsx export. Install it with: uv add openpyxl"
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Export"

        # header row
        for col, name in enumerate(field_names, 1):
            ws.cell(row=1, column=col, value=name)

        # data rows
        for row_idx, row in enumerate(self._iter_rows(qs, field_names, total), 2):
            for col, name in enumerate(field_names, 1):
                val = row[name]
                ws.cell(row=row_idx, column=col, value=str(val) if val != "" else None)

        wb.save(output_path)
